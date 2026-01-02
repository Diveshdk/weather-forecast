"""
IMD Heavy Rainfall Verification System
Module: Professional Report Generator (Excel & PDF)
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
from typing import Dict, List, Optional
import os
import sys

# Add the app directory to Python path for imports
current_dir = os.path.dirname(os.path.abspath(__file__        dist_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])) = os.path.dirname(current_dir)
sys.path.append(app_dir)

from utils.heavyRainfallVerifier import HeavyRainfallVerifier, SkillScores, DistrictVerificationSummary
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

class IMDReportGenerator:
    """Generate comprehensive Excel and PDF reports for Heavy Rainfall Verification"""
    
    def __init__(self, verification_results: Dict):
        """
        Initialize report generator with verification results
        
        Args:
            verification_results: Output from HeavyRainfallVerifier.run_full_verification()
        """
        self.results = verification_results
        self.threshold = verification_results['heavy_threshold']
        self.month = verification_results['target_month']
        self.districts = verification_results['districts_analyzed']
        self.skill_scores = verification_results['skill_scores']
        self.district_summaries = verification_results['district_summaries']
        
        # Set up matplotlib style for charts
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
    
    def generate_excel_report(self, output_path: str) -> str:
        """
        Generate comprehensive Excel report with multiple sheets
        
        Args:
            output_path: Path for the Excel file
        
        Returns:
            Full path of generated Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_skill_scores_sheet(wb)
        self._create_district_results_sheet(wb)
        self._create_contingency_tables_sheet(wb)
        self._create_verification_details_sheet(wb)
        
        # Save workbook
        full_path = os.path.abspath(output_path)
        wb.save(full_path)
        print(f"Excel report saved: {full_path}")
        
        return full_path
    
    def _create_summary_sheet(self, wb: Workbook):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title and header
        title_font = Font(size=16, bold=True, color="2F5597")
        header_font = Font(size=12, bold=True)
        
        ws['A1'] = f"IMD Heavy Rainfall Verification Report - {self.month} 2025"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        
        # Summary statistics
        ws['A3'] = "Verification Summary"
        ws['A3'].font = header_font
        
        summary_data = [
            ['Heavy Rainfall Threshold', f'{self.threshold} mm'],
            ['Total Verifications', self.results['total_verifications']],
            ['Districts Analyzed', len(self.districts)],
            ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for i, (label, value) in enumerate(summary_data, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Overall performance by lead time
        ws['A9'] = "Overall Performance by Lead Time"
        ws['A9'].font = header_font
        
        # Headers
        headers = ['Lead Time', 'Hits', 'Misses', 'False Alarms', 'Correct Negatives', 'POD', 'FAR', 'CSI', 'Bias']
        for i, header in enumerate(headers, start=1):
            ws.cell(row=10, column=i, value=header).font = Font(bold=True)
        
        # Data
        for i, (lead_key, scores) in enumerate(self.skill_scores.items(), start=11):
            ws.cell(row=i, column=1, value=lead_key)
            ws.cell(row=i, column=2, value=scores.hits)
            ws.cell(row=i, column=3, value=scores.misses)
            ws.cell(row=i, column=4, value=scores.false_alarms)
            ws.cell(row=i, column=5, value=scores.correct_negatives)
            ws.cell(row=i, column=6, value=round(scores.pod, 3))
            ws.cell(row=i, column=7, value=round(scores.far, 3))
            ws.cell(row=i, column=8, value=round(scores.csi, 3))
            ws.cell(row=i, column=9, value=round(scores.bias, 3))
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_skill_scores_sheet(self, wb: Workbook):
        """Create detailed skill scores sheet"""
        ws = wb.create_sheet("Skill Scores")
        
        # Create DataFrame for skill scores
        scores_data = []
        for lead_key, scores in self.skill_scores.items():
            scores_data.append({
                'Lead Time': lead_key,
                'Hits': scores.hits,
                'Misses': scores.misses,
                'False Alarms': scores.false_alarms,
                'Correct Negatives': scores.correct_negatives,
                'Total': scores.total,
                'POD (Probability of Detection)': round(scores.pod, 4),
                'FAR (False Alarm Ratio)': round(scores.far, 4),
                'CSI (Critical Success Index)': round(scores.csi, 4),
                'Bias (Frequency Bias)': round(scores.bias, 4),
                'TSS (True Skill Statistic)': round(scores.tss, 4),
                'HSS (Heidke Skill Score)': round(scores.hss, 4)
            })
        
        df = pd.DataFrame(scores_data)
        
        # Add to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    def _create_district_results_sheet(self, wb: Workbook):
        """Create district-wise results sheet"""
        ws = wb.create_sheet("District Results")
        
        # Create comprehensive district results table
        district_data = []
        for district in self.districts:
            for lead_key in ['Day-1', 'Day-2', 'Day-3', 'Day-4', 'Day-5']:
                if lead_key in self.district_summaries[district]:
                    summary = self.district_summaries[district][lead_key]
                    district_data.append({
                        'District': district,
                        'Lead Time': lead_key,
                        'Hits': summary.hits,
                        'Misses': summary.misses,
                        'False Alarms': summary.false_alarms,
                        'Correct Negatives': summary.correct_negatives,
                        'Total': summary.total,
                        'Accuracy %': round(summary.accuracy_percent, 1),
                        'POD': round(summary.pod, 3),
                        'FAR': round(summary.far, 3),
                        'CSI': round(summary.csi, 3),
                        'Bias': round(summary.bias, 3)
                    })
        
        df = pd.DataFrame(district_data)
        
        # Add to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Apply conditional formatting for accuracy
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    if cell.value >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif cell.value >= 60:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def _create_contingency_tables_sheet(self, wb: Workbook):
        """Create contingency tables for each lead time"""
        ws = wb.create_sheet("Contingency Tables")
        
        row = 1
        for lead_key, scores in self.skill_scores.items():
            # Title
            ws.cell(row=row, column=1, value=f"Contingency Table - {lead_key}").font = Font(bold=True)
            row += 2
            
            # Create contingency table
            contingency_data = [
                ['', 'Observed Heavy', 'Observed No Heavy', 'Total'],
                ['Forecast Heavy', scores.hits, scores.false_alarms, scores.hits + scores.false_alarms],
                ['Forecast No Heavy', scores.misses, scores.correct_negatives, scores.misses + scores.correct_negatives],
                ['Total', scores.hits + scores.misses, scores.false_alarms + scores.correct_negatives, scores.total]
            ]
            
            for r_data in contingency_data:
                for c_idx, value in enumerate(r_data, start=1):
                    cell = ws.cell(row=row, column=c_idx, value=value)
                    if row == row:  # Header row
                        cell.font = Font(bold=True)
                row += 1
            
            row += 2  # Space between tables
    
    def _create_verification_details_sheet(self, wb: Workbook):
        """Create detailed verification results sheet (sample data)"""
        ws = wb.create_sheet("Verification Details")
        
        # Sample of verification results (first 100 to avoid huge file)
        sample_results = self.results['verification_results'][:100]
        
        details_data = []
        for result in sample_results:
            details_data.append({
                'District': result.district,
                'Date': result.date,
                'Lead Days': f'Day-{result.lead_days}',
                'Forecast Code': result.forecast_code,
                'Forecast Heavy': 'Yes' if result.forecast_heavy else 'No',
                'Observed Rainfall (mm)': round(result.observed_rainfall, 1),
                'Observed Heavy': 'Yes' if result.observed_heavy else 'No',
                'Result': result.result_type,
                'Description': result.result_description
            })
        
        df = pd.DataFrame(details_data)
        
        # Add to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    def generate_pdf_report(self, output_path: str) -> str:
        """
        Generate professional PDF report
        
        Args:
            output_path: Path for the PDF file
        
        Returns:
            Full path of generated PDF file
        """
        full_path = os.path.abspath(output_path)
        doc = SimpleDocTemplate(full_path, pagesize=A4)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        elements = []
        
        # Title
        title = Paragraph(f"IMD Heavy Rainfall Verification Report<br/>{self.month} 2025", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Executive Summary
        elements.append(Paragraph("Executive Summary", styles['Heading2']))
        
        summary_text = f"""
        This report presents the verification results for heavy rainfall forecasts (≥{self.threshold}mm) 
        for {self.month} 2025. A total of {self.results['total_verifications']} forecast-observation pairs 
        were analyzed across {len(self.districts)} districts for lead times from Day-1 to Day-5.
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Overall Performance Table
        elements.append(Paragraph("Overall Performance by Lead Time", styles['Heading3']))
        
        # Create performance table data
        perf_data = [['Lead Time', 'POD', 'FAR', 'CSI', 'Bias', 'Total Cases']]
        for lead_key, scores in self.skill_scores.items():
            perf_data.append([
                lead_key,
                f"{scores.pod:.3f}",
                f"{scores.far:.3f}",
                f"{scores.csi:.3f}",
                f"{scores.bias:.3f}",
                str(scores.total)
            ])
        
        perf_table = Table(perf_data)
        perf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(perf_table)
        elements.append(PageBreak())
        
        # Key Findings
        elements.append(Paragraph("Key Findings", styles['Heading2']))
        
        # Find best and worst performing lead times
        best_csi_lead = max(self.skill_scores.keys(), key=lambda x: self.skill_scores[x].csi)
        best_pod_lead = max(self.skill_scores.keys(), key=lambda x: self.skill_scores[x].pod)
        
        findings_text = f"""
        <b>Performance Highlights:</b><br/>
        • Best CSI (Critical Success Index): {best_csi_lead} with {self.skill_scores[best_csi_lead].csi:.3f}<br/>
        • Best POD (Probability of Detection): {best_pod_lead} with {self.skill_scores[best_pod_lead].pod:.3f}<br/>
        • Overall bias shows {"overforecasting" if self.skill_scores['Day-1'].bias > 1 else "underforecasting"} tendency<br/>
        • False Alarm Ratio ranges from {min(scores.far for scores in self.skill_scores.values()):.3f} to {max(scores.far for scores in self.skill_scores.values()):.3f}
        """
        
        elements.append(Paragraph(findings_text, styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # District Performance Summary (top 10)
        elements.append(Paragraph("Top Performing Districts (Day-1 Accuracy)", styles['Heading3']))
        
        # Get top districts by Day-1 accuracy
        day1_districts = [(district, summary['Day-1'].accuracy_percent) 
                         for district, summary in self.district_summaries.items() 
                         if 'Day-1' in summary]
        day1_districts.sort(key=lambda x: x[1], reverse=True)
        
        dist_data = [['District', 'Accuracy (%)', 'POD', 'CSI']]
        for district, accuracy in day1_districts[:10]:
            day1_summary = self.district_summaries[district]['Day-1']
            dist_data.append([
                district,
                f"{accuracy:.1f}",
                f"{day1_summary.pod:.3f}",
                f"{day1_summary.csi:.3f}"
            ])
        
        dist_table = Table(dist_data)
        dist_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), reportlab_colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), reportlab_colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), reportlab_colors.white),
            ('GRID', (0, 0), (-1, -1), 1, reportlab_colors.black)
        ]))
        
        elements.append(dist_table)
        
        # Build PDF
        doc.build(elements)
        print(f"PDF report saved: {full_path}")
        
        return full_path
    
    def generate_charts(self, output_dir: str) -> List[str]:
        """
        Generate visualization charts and save as image files
        
        Args:
            output_dir: Directory to save chart images
        
        Returns:
            List of paths to generated chart files
        """
        os.makedirs(output_dir, exist_ok=True)
        chart_files = []
        
        # Chart 1: Skill Scores by Lead Time
        plt.figure(figsize=(12, 8))
        leads = list(self.skill_scores.keys())
        pod_scores = [self.skill_scores[lead].pod for lead in leads]
        far_scores = [self.skill_scores[lead].far for lead in leads]
        csi_scores = [self.skill_scores[lead].csi for lead in leads]
        
        x = np.arange(len(leads))
        width = 0.25
        
        plt.bar(x - width, pod_scores, width, label='POD', alpha=0.8)
        plt.bar(x, far_scores, width, label='FAR', alpha=0.8)
        plt.bar(x + width, csi_scores, width, label='CSI', alpha=0.8)
        
        plt.xlabel('Lead Time')
        plt.ylabel('Score')
        plt.title(f'Heavy Rainfall Verification Skill Scores - {self.month} 2025')
        plt.xticks(x, leads)
        plt.legend()
        plt.grid(True, alpha=0.3)
        
        chart1_path = os.path.join(output_dir, 'skill_scores_by_lead_time.png')
        plt.savefig(chart1_path, dpi=300, bbox_inches='tight')
        plt.close()
        chart_files.append(chart1_path)
        
        # Chart 2: District Accuracy Comparison (Day-1)
        plt.figure(figsize=(14, 8))
        district_accuracies = [(district, summary['Day-1'].accuracy_percent) 
                              for district, summary in self.district_summaries.items() 
                              if 'Day-1' in summary]
        district_accuracies.sort(key=lambda x: x[1], reverse=True)
        
        districts, accuracies = zip(*district_accuracies[:15])  # Top 15
        
        plt.bar(range(len(districts)), accuracies, alpha=0.8)
        plt.xlabel('District')
        plt.ylabel('Accuracy (%)')
        plt.title(f'District Accuracy Comparison (Day-1) - {self.month} 2025')
        plt.xticks(range(len(districts)), districts, rotation=45, ha='right')
        plt.grid(True, alpha=0.3)
        
        chart2_path = os.path.join(output_dir, 'district_accuracy_day1.png')
        plt.savefig(chart2_path, dpi=300, bbox_inches='tight')
        plt.close()
        chart_files.append(chart2_path)
        
        return chart_files

def generate_full_report(files_dir: str, output_dir: str, target_month: str = "June", 
                        heavy_threshold: float = 64.5) -> Dict[str, str]:
    """
    Generate complete verification report with Excel, PDF, and charts
    
    Args:
        files_dir: Directory containing IMD data files
        output_dir: Directory to save reports
        target_month: Month to analyze
        heavy_threshold: Heavy rainfall threshold in mm
    
    Returns:
        Dictionary with paths to generated files
    """
    print(f"Starting full report generation for {target_month}...")
    
    # Run verification
    verifier = HeavyRainfallVerifier(heavy_threshold=heavy_threshold)
    results = verifier.run_full_verification(files_dir, target_month)
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Initialize report generator
    generator = IMDReportGenerator(results)
    
    # Generate reports
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Excel report
    excel_path = os.path.join(output_dir, f'IMD_Heavy_Rainfall_Verification_{target_month}_{timestamp}.xlsx')
    excel_file = generator.generate_excel_report(excel_path)
    
    # PDF report
    pdf_path = os.path.join(output_dir, f'IMD_Heavy_Rainfall_Verification_{target_month}_{timestamp}.pdf')
    pdf_file = generator.generate_pdf_report(pdf_path)
    
    # Charts
    charts_dir = os.path.join(output_dir, f'charts_{timestamp}')
    chart_files = generator.generate_charts(charts_dir)
    
    generated_files = {
        'excel_report': excel_file,
        'pdf_report': pdf_file,
        'charts': chart_files,
        'charts_directory': charts_dir
    }
    
    print(f"\nReport generation completed!")
    print(f"Excel Report: {excel_file}")
    print(f"PDF Report: {pdf_file}")
    print(f"Charts: {len(chart_files)} files in {charts_dir}")
    
    return generated_files

if __name__ == "__main__":
    # Test report generation
    files_dir = "/Users/divesh/Desktop/imdmumbai/imdfiles"
    output_dir = "/Users/divesh/Desktop/imdmumbai/reports"
    
    generated_files = generate_full_report(files_dir, output_dir, "June", 64.5)
    print("\nGenerated files:", generated_files)
